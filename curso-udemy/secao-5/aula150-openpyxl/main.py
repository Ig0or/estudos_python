import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

for campo in planilha1['b']:
    if campo.value:
        print(campo.value)
print()

for linha in planilha1['a1:c2']:
    for coluna in linha:
        print(coluna.value)
print()

planilha1['b3'].value = 5000

for linha in planilha1:
    if linha[0].value:
        print(linha[0].value, end=' ')
    if linha[1].value:
        print(linha[1].value, end=' ')
    if linha[2].value:
        print(linha[2].value)

for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')
